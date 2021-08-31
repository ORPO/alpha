import importlib
import uuid
import shutil

import openpyxl
from lxml import etree

from sql_models.models import *


def readuso():  # 280521
    Uso.create_table()
    try:
        wb = openpyxl.load_workbook("E:\Sosedka\\io_.xlsx")
        for sheet in wb.worksheets:
            if sheet.title == 'HW':
                print("2")
                dt = []
                a_dict = []
                rows = sheet.max_row
                col = sheet.max_column
                for i in range(4, rows + 1):
                    for j in range(7,col+1):

                        if str(sheet.cell(row=i, column=j).value)[:1] == 'm':


                            a_dict = dict(alpha=(str(sheet.cell(row=i, column=3).value)+str(sheet.cell(row=2, column=j).value)),
                                          opc_ua=('D_'+str(sheet.cell(row=i, column=1).value)+str(sheet.cell(row=2, column=j).value)))
                        dt.append(a_dict)
                        print(a_dict)



    except:
        print("не записано")

readuso()