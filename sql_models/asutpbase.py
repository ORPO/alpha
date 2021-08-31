import openpyxl
from lxml import etree
from peewee import *
import markdown
import time

from models import *

format_analog = {'I': '1',
                 'T': '1',
                 'X': '1',
                 'Газ': '1',
                 'Пожар': '1',
                 'P': '3',
                 'dP': '3',
                 'L': '0'}
db = PostgresqlDatabase('asutp', user='postgres', password='P@ssw0rd2021',
                        host='10.157.20.34', port=5432)

def base_write_many2(objtables, data):
    try:
        for idx in range(0, len(data), 500):
            objtables.insert_many(data[idx:idx + 500]).on_conflict('replace').execute()
        print("обновлено", objtables)
    except:
        print('обновление не удалась', objtables)
    try:
        print(len(data))
        for idx in range(0, len(data), 500):
            objtables.insert_many(data[idx:idx + 500]).on_conflict('ignore').execute()
        print("запись завершена, записано", objtables)
    except:
        print("не записано", objtables)

def str_find_int(str1, arr):
    i=0
    for el in arr:
        if str(str1).find(el) > -1:
           i=i+1
    return i

sch = "setpoints"
class Basemodel(Model):  # базовый клас
    class Meta:
        database = db  # модель будет использовать базу данных указанную выше

class  layout(Basemodel):
        base = IntegerField(null=True)
        mult = IntegerField(null=True)
        mult1 = IntegerField(null=True)
        mult2 = IntegerField(null=True)
        mult3 = IntegerField(null=True)
        start_table =CharField(null=True)
        caption = CharField(null=True)
        msgsstart = IntegerField(null=True)
        msgsend = IntegerField(null=True)
        msgscount = IntegerField(null=True)
        id = IntegerField(primary_key=True,default="nextval('messages.layout_id_seq'::regclass)")
        class Meta:
            schema = sch
# for obj in layout.select():
#     print(obj)

class tblanaloggroups(Basemodel):
    id = IntegerField(primary_key=True)
    prefix = CharField(null=True)
    name= CharField(null=True)
    min6name= CharField(null=True)
    min5name= CharField(null=True)
    min4name= CharField(null=True)
    min3name= CharField(null=True)
    min2name= CharField(null=True)
    min1name= CharField(null=True)
    max1name= CharField(null=True)
    max2name= CharField(null=True)
    max3name= CharField(null=True)
    max4name= CharField(null=True)
    max5name= CharField(null=True)
    max6name= CharField(null=True)
    messagetable= CharField(null=True)
    hintskey= CharField(null=True)
    class Meta:
        schema = sch

class tblanalogs(Basemodel):
    id = IntegerField(primary_key=True)
    prefix = CharField(null=True)
    systemindex = IntegerField(default=0)
    tag= CharField(null=True)
    name= CharField(null=True)
    analoggroupid = IntegerField(null=True,default=1)
    setpointgroupid = IntegerField(null=True,default=1)
    egu= CharField(null=True)
    isoilpressure = BooleanField(default=False)
    isinterface  = BooleanField(default=False)
    isphysic = BooleanField(default=False)
    ispumpvibration=BooleanField(null=True)
    precision = IntegerField(default=1)
    istrending= BooleanField(default=True)
    trendingsettings=CharField(null=True)
    trendinggroup=IntegerField(null=True)
    lolimfield=FloatField(null=True)
    hilimfield=FloatField(null=True)
    lolimeng=FloatField(null=True)
    hilimeng=FloatField(null=True)
    lolim=FloatField(null=True)
    hilim=FloatField(null=True)
    min6=FloatField(null=True)
    min5=FloatField(null=True)
    min4=FloatField(null=True)
    min3=FloatField(null=True)
    min2=FloatField(null=True)
    min1=FloatField(null=True)
    max1=FloatField(null=True)
    max2=FloatField(null=True)
    max3=FloatField(null=True)
    max4=FloatField(null=True)
    max5=FloatField(null=True)
    max6=FloatField(null=True)
    histeresis=FloatField(null=True)
    deltahi=FloatField(null=True)
    deltalo=FloatField(null=True)
    deltat=FloatField(null=True)
    smoothfactor=FloatField(null=True)
    ctrl=SmallIntegerField(null=True)
    msgmask=IntegerField(null=True)
    sigmask=IntegerField(null=True)
    ctrlmask=SmallIntegerField(null=True)
    timefilter = FloatField(null=True)
    isbackup=BooleanField(default=False)
    tabindex=SmallIntegerField(null=True)
    rulename=CharField(null=True)
    hintskey=CharField(null=True)
    class Meta:
        schema = sch


def analogsgrp():
    wb = openpyxl.load_workbook("E:\тестовый\опытный\\analogs.xlsx")
    for sheet in wb.worksheets:
        if sheet.title == 'tagrp':
            data = []
            a_dict = []
            rows = sheet.max_row
            for i in range(2, rows + 1):
                tblanaloggroups.insert(id = sheet.cell(row=i, column=1).value,
                                prefix = sheet.cell(row=i, column=2).value,
                                name= sheet.cell(row=i, column=3).value,
                                min6name= sheet.cell(row=i, column=4).value,
                                min5name= sheet.cell(row=i, column=5).value,
                                min4name= sheet.cell(row=i, column=6).value,
                                min3name= sheet.cell(row=i, column=7).value,
                                min2name= sheet.cell(row=i, column=8).value,
                                min1name= sheet.cell(row=i, column=9).value,
                                max1name= sheet.cell(row=i, column=10).value,
                                max2name= sheet.cell(row=i, column=11).value,
                                max3name= sheet.cell(row=i, column=12).value,
                                max4name= sheet.cell(row=i, column=13).value,
                                max5name= sheet.cell(row=i, column=14).value,
                                max6name= sheet.cell(row=i, column=15).value,
                                messagetable= sheet.cell(row=i, column=16).value,
                                hintskey= sheet.cell(row=i, column=17).value).execute()




def tbl_analogs():
    tabindex=1#разобраться что это
    ruleindex=1
    isoilpressure = False
    AnalogGroupId=1
    lolimfield=4000
    hilimfield=20000
    lolim=3900
    hilim=20100
    lolimeng=4
    hilimeng=20
    physicegu='мкА'
    isphysic=True
    ispumpvibration=False
    for obj in Analog.select():
        a_dict=[]
        b_dict=[]
        if str_find_int(obj.description.lower(), ['ос', 'смещ']) >= 2:
            AnalogGroupId=8
        if str_find_int(obj.description.lower(), ['темп']) >= 1:
            AnalogGroupId=2
            lolimeng=-50
            hilimeng=150
        if str_find_int(obj.description.lower(), ['темп','масл']) >= 2:
            AnalogGroupId=3
        if str_find_int(obj.description.lower(), ['темп','двиг']) >= 2:
            AnalogGroupId=4
        if str_find_int(obj.description.lower(), ['урове']) >= 1:
            AnalogGroupId=5
        if str_find_int(obj.description.lower(), ['вибра','ЭД']) >= 2:
            AnalogGroupId=6
        if str_find_int(obj.description.lower(), ['вибра','насо']) >= 2:
            AnalogGroupId=7
            ispumpvibration=True
        if str_find_int(obj.description.lower(), ['давл']) >= 1:
            AnalogGroupId=9
            lolimeng=0
            hilimeng=1.02
        if str_find_int(obj.description.lower(), ['пере','давл']) >= 2:
            AnalogGroupId=10
            lolimeng=0
            hilimeng=16
        if str_find_int(obj.description.lower(), ['загаз']) >= 1:
            AnalogGroupId=11
        if str_find_int(obj.description.lower(), ['уров','утеч']) >= 2:
            AnalogGroupId=12
        if str_find_int(obj.description.lower(), ['уров','масл']) >= 2:
            AnalogGroupId=13
        if str_find_int(obj.description.lower(), ['вибра','ЭД']) >= 2:
            AnalogGroupId=15

        if str_find_int(obj.description.lower(), ['давл','нефт']) >= 2:
            isoilpressure=True

        if str_find_int(obj.description.lower(), ['вибр']) >= 1:
            ruleindex=5
            lolimeng=0
            hilimeng=30

        a_dict=dict(
                          tag=obj.tag,
                          name=obj.description,
                          egu=obj.egu,
                          precision=format_analog[obj.sign],
                          analoggroupid=AnalogGroupId,
                          isoilpressure=isoilpressure,
                          tabindex=tabindex,
                          rulename=ruleindex,
                          lolimfield=lolimfield,
                          hilimfield = hilimfield,
                          lolim = lolim,
                          hilim = hilim,
                          lolimeng = lolimeng,
                          hilimeng = hilimeng,
                          ispumpvibration=ispumpvibration,
                          isphysic=isphysic)

        try:
            b_dict = a_dict
            b_dict["id"]=obj.id
            tblanalogs.insert(b_dict).execute()
        except:
            tblanalogs.update(a_dict).where(tblanalogs.tag==obj.tag).execute()
#tbl_analogs()

def ustanalogs():
    wb = openpyxl.load_workbook("E:\тестовый\опытный\\analogs.xlsx")
    for sheet in wb.worksheets:
        if sheet.title == '1':
            data = []
            a_dict = []
            rows = sheet.max_row
            for i in range(2, rows + 1):
                print(str(sheet.cell(row=i, column=1).value).replace('.','_'))
                for obj in tblanalogs.select().where(tblanalogs.tag==(str(sheet.cell(row=i, column=1).value).replace('.','_'))):
                    #print(obj.tag)
                    tblanalogs.update(lolimeng=sheet.cell(row=i, column=4).value,
                                      hilimeng=sheet.cell(row=i, column=5).value).where(tblanalogs.id==obj.id).execute()

ustanalogs()
