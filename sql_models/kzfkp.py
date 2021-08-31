import openpyxl


xls_sheme = dict(typesignal='Тип сигнала',
                cabinet='Шкаф',
                tag='Тэг',
                desc='Наименование',
                sheme='Схема',
                termblock='КлК',
                contacts='Конт',
                unit='Корз',
                module='Мод',
                channel='Кан'
                )

def get_key(d, value):
    for k, v in d.items():
        if v == value:
            return k

xls_path = 'E:\sandbox\\kzfkp.xlsx'


def read_xls_sheme(xls_path,xls_sheme,xls_list):
    wb = openpyxl.load_workbook(xls_path)

    for sheet in wb.worksheets:
        if sheet.title == xls_list:
            rows = sheet.max_row
            cols = sheet.max_column
            keys = []
            values = []
            for i in range(1, rows + 1):
                for j in range(1,cols+1):
                    if sheet.cell(row=i,column=j).value in xls_sheme.values():
                        start = i
                        keys.append('start_row')
                        values.append(i)
                        break

            for i in range(1, cols+1):
                if sheet.cell(row=start,column=i).value in xls_sheme.values():#             for i in range (1,50):
                    val = sheet.cell(row=start,column=i).value
                    key = get_key(xls_sheme,val)
                    keys.append(key)
                    values.append(i)


            x={k:v for k, v in zip(keys, values)}
            ret_list={xls_list:x}
            return ret_list
a=[]
for list in ('МНС3.КЦ','МНС3.УСО.1(1)'):
    a.append(read_xls_sheme(xls_path,xls_sheme,list))
print(a)