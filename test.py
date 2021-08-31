import importlib
import uuid
import shutil

import openpyxl
from lxml import etree

from sql_models.models import *

TypeSignal = ['AI', 'AO', 'DI', 'DO']
CabinetDict = {'МНС.САР': 'SAR_KC', 'МНС.КЦ': 'MNS_KC', 'МНС.УСО.1(1) с БРУ': 'USO_1_1', 'МНС.УСО.1(2)': 'USO_1_2',
               'МНС.УСО.1(3)': 'USO_1_3', 'МНС.УСО.2': 'USO_2', 'МНС.УСО.3': 'USO_3', 'МНС.УСО.4': 'USO_4'}

xlslists = ['МНС.САР', 'МНС.КЦ', 'МНС.УСО.1(1) с БРУ', 'МНС.УСО.1(2)', 'МНС.УСО.1(3)', 'МНС.УСО.2', 'МНС.УСО.3',
            'МНС.УСО.4']
channel = ['1','2','3','4','5','6','7','8','9']
modId = {'521':'DI', '531': 'DO', '514':'AO', '516':'AI','502':'CPU','503':'CPU','541':'RS','546':'MN','545':'CN','550':'PSU'}


# class kzfkp(Signal):
#     class Meta:
#         database = db  # модель будет использовать базу данных указанную выше

def str_find(str1, arr):
    i=0
    for el in arr:
        if str(str1).find(el) > -1:
            return True

def str_find_int(str1, arr):
    i=0
    for el in arr:
        if str(str1).find(el) > -1:
           i=i+1
    return i


def readkzfkp():
    Kzfkp.create_table()
    try:
        data = []
        a_dict = []

        wb = openpyxl.load_workbook("E:\sandbox\MK500\\kzfkp.xlsx")
        for sheet in wb.worksheets:
            qr = Uso.select().where(Uso.description == str(sheet.title))
            if qr.exists():
                print('лист найден')
            if sheet.title in xlslists:
                rows = sheet.max_row
                for i in range(14, rows + 1):
                    if sheet.cell(row=i, column=11).value is not None and sheet.cell(row=i, column=2).value is not None:
                        ch = str(sheet.cell(row=i, column=13).value)
                        if str(sheet.cell(row=i, column=13).value) in channel:
                            ch = '0'+ str(sheet.cell(row=i, column=13).value)
                        Signal_id = str(CabinetDict[sheet.cell(row=i, column=2).value]) + '_A' + str(
                            sheet.cell(row=i, column=11).value) + \
                                    '_0' + str(sheet.cell(row=i, column=12).value) + '_' + ch
                        print(Signal_id)
                        CabinetTag = str(CabinetDict[sheet.cell(row=i, column=2).value])


                        USO_id = str(CabinetDict[sheet.cell(row=i, column=2).value]) + '_A' + str(
                            sheet.cell(row=i, column=11).value) + \
                                 '_0' + str(sheet.cell(row=i, column=12).value)

                        module = importlib.import_module('sql_models.models')
                        table = getattr(module, 'UsoModule')


                        qr = UsoModule.select().where(UsoModule.uso_module_id == USO_id)
                        HWDesc = ''
                        if qr.exists():
                            obj = qr.get()
                            HWDesc = obj.module_type_id


                        a_dict = dict(cabinetName=sheet.cell(row=i, column=2).value,
                                      tag=sheet.cell(row=i, column=3).value,
                                      signalname=sheet.cell(row=i, column=4).value,
                                      sheme=sheet.cell(row=i, column=5).value,
                                      terminalblock=sheet.cell(row=i, column=7).value,
                                      kont=sheet.cell(row=i, column=8).value,
                                      unit=sheet.cell(row=i, column=11).value,
                                      hwdesc=HWDesc,
                                      cabinettag=CabinetTag,
                                      module=sheet.cell(row=i, column=12).value,
                                      channel=ch,
                                      typesignal=sheet.cell(row=i, column=15).value,
                                      signal_id=Signal_id,
                                      typesheme=str(sheet.cell(row=i, column=5).value).replace(" ", ""),
                                      ttips=sheet.cell(row=i, column=14).value)
                        data.append(a_dict)
        base_write_many2(Kzfkp, data)
        print("записано")

    except:
        print("не записано")


def readuso():  # 280521
    Uso.create_table()
    try:
        wb = openpyxl.load_workbook("E:\sandbox\MK500\module.xlsx")
        for sheet in wb.worksheets:
            if sheet.title == 'USO':
                data = []
                a_dict = []
                rows = sheet.max_row
                for i in range(2, rows + 1):
                    a_dict = dict(uso_id=sheet.cell(row=i, column=2).value,
                                  description=sheet.cell(row=i, column=1).value)
                    data.append(a_dict)
                base_write_many2(Uso, data)


    except:
        print("не записано")


def readtypemodule():  # 280521
    ModuleType.create_table()
    try:
        wb = openpyxl.load_workbook("E:\sandbox\MK500\module.xlsx")
        for sheet in wb.worksheets:
            if sheet.title == 'Type_module':
                data = []
                a_dict = []
                rows = sheet.max_row
                for i in range(2, rows + 1):
                    a_dict = dict(module_type_id=str(sheet.cell(row=i, column=1).value)[3:6],
                                  sign=sheet.cell(row=i, column=1).value,
                                  description=sheet.cell(row=i, column=2).value,
                                  tooltips=sheet.cell(row=i, column=3).value,
                                  hwdesc=sheet.cell(row=i, column=4).value,
                                  quantity=int(sheet.cell(row=i, column=5).value))

                    data.append(a_dict)
                base_write_many2(ModuleType, data)
    except:
        print("не записано")


def readmodule():  # 280521
    UsoModule.create_table()
    try:
        wb = openpyxl.load_workbook("E:\sandbox\MK500\module.xlsx")
        for sheet in wb.worksheets:
            if sheet.title == 'USO_module':
                data = []
                a_dict = []
                rows = sheet.max_row
                print(rows)
                for i in range(2, rows + 1):
                    qr = Uso.select().where(Uso.description == sheet.cell(row=i, column=1).value)
                    print(qr.get())
                    a_dict = dict(uso_id=qr.get(),
                                  position=sheet.cell(row=i, column=2).value,
                                  sign=sheet.cell(row=i, column=3).value,
                                  module_type_id=sheet.cell(row=i, column=4).value[3:6],
                                  uso_module_id=str(qr.get()) + "_" + str(sheet.cell(row=i, column=2).value),
                                  index_arr = sheet.cell(row=i, column=5).value)
                    print(str(qr.get()) + "_" + str(sheet.cell(row=i, column=2).value))
                    data.append(a_dict)
                base_write_many2(UsoModule, data)


    except:
        print("не записано")


# def test_module_sign():
#     for obj in kzfkp.select(kzfkp,Module_type)\
#             .switch(kzfkp)\
#             .join(Module_type, on=(kzfkp.HWDesc == Module_type.Module_type_id))\
#             .order_by(kzfkp.CabinetName,kzfkp.HWDesc):
#         print(obj.CabinetName, obj.Tag, obj.SignalName)
#
# def atr_desc_():
#     for obj in USO_module.select():
#         Mtype=obj.Module_type_id
#         el=object_atributes.select().where(object_atributes.object_id == Mtype)
#         for idx in range(1,17):
#             if el.exists():
#                 atr = getattr(el.get(), ('atr'+str(idx)))
#                 if atr is not None:
#                     qr = Prj_attributes.select().where(Prj_attributes.Attributes_id==atr)
#                     if qr.exists():
#                         val = getattr(qr.get(),'Attributes_value')
#                         tbl, dot, col = val.rpartition('.')
#                         if tbl == '':
#                             try:
#                                t1 = Module_type.select().where(Module_type.Module_type_id == Mtype)#.get()
#
#                                get_atr(Module_type,col,Mtype,";")
#                             except:
#                                 print('2')
#                         # else:
#                         #     module = importlib.import_module('sql_models.models')
#                         #     # table = getattr(module, tbl,Mtype,";")
#                         #     #el=table.select()
#                         #     #     el1 = getattr(el, col)
#                         #     #     print(el1)

def get_atr(curr_table, source, item, delimiter):
    ret_atr = []

    source_dict = str(source).split(';')
    str_tbl = str(curr_table).replace("<Model: ", "").replace(">", "")
    for src in source_dict:
        tbl, dot, col = src.rpartition('.')
        print(tbl)
        if tbl == '':
            try:
                module = importlib.import_module('sql_models.models')
                table = getattr(module, str_tbl)
                str_id = str(str_tbl) + '_id'
                sid = getattr(table, str_id)
                qr = table.select().where(sid == item).get()
                ret_atr.append(getattr(qr, col))
            except:
                print('1')
        else:
            try:
                module = importlib.import_module('sql_models.models')
                table = getattr(module, tbl)
                str_tbl = str(table).replace("<Model: ", "").replace(">", "")
                str_id = str(str_tbl) + '_id'
                sid = getattr(table, str_id)
                print(sid)
                qr = table.select().where(sid == item).get()
                ret_atr.append(getattr(qr, col))
            except:
                print('2')

    return delimiter.join(ret_atr)


def add_module(filename_omx):
    try:
        print(filename_omx)
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(filename_omx, parser)
        root = tree.getroot()
        for el in root.iter('{automation.deployment}application-object'):
            if el.attrib['name'] == "Application_PLC":
                apl = el
                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'DIs':
                        fold = el1
                        for obj in UsoModule.select().where(UsoModule.module_type_id == '521'):
                            mtype = ModuleType.select().where(ModuleType.module_type_id == '521').get()
                            usoname = Uso.select().where(Uso.uso_id == obj.uso_id).get()
                            name = obj.uso_module_id
                            for el in fold.iter('{automation.control}object'):
                                if el.attrib['name'] == obj.uso_module_id:
                                    fold.remove(el)
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = name
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.modules.mod_DI"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.ModPosition"
                            atrb1.attrib['value'] = str(obj.sign)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.System.Attributes.Description"
                            atrb2.attrib['value'] = mtype.sign
                            object.append(atrb2)
                            atrb3 = etree.Element("attribute")
                            atrb3.attrib['type'] = "unit.Library.Attributes.ModID"
                            atrb3.attrib['value'] = mtype.hwdesc
                            object.append(atrb3)
                            atrb4 = etree.Element("attribute")
                            atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                            atrb4.attrib['value'] = usoname.description
                            object.append(atrb4)
                            fold.append(object)
                            tree.write(filename_omx,pretty_print=True)

                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'AOs':
                        fold = el1
                        for obj in UsoModule.select().where(UsoModule.module_type_id == '514'):
                            mtype = ModuleType.select().where(ModuleType.module_type_id == '514').get()
                            usoname = Uso.select().where(Uso.uso_id == obj.uso_id).get()
                            for el in fold.iter('{automation.control}object'):
                                if el.attrib['name'] == obj.uso_module_id:
                                    fold.remove(el)
                            name = obj.uso_module_id
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = name
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.modules.mod_AI"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.ModPosition"
                            atrb1.attrib['value'] = str(obj.sign)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.System.Attributes.Description"
                            atrb2.attrib['value'] = mtype.sign
                            object.append(atrb2)
                            atrb3 = etree.Element("attribute")
                            atrb3.attrib['type'] = "unit.Library.Attributes.ModID"
                            atrb3.attrib['value'] = mtype.hwdesc
                            object.append(atrb3)
                            atrb4 = etree.Element("attribute")
                            atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                            atrb4.attrib['value'] = usoname.description
                            object.append(atrb4)
                            fold.append(object)
                            tree.write(filename_omx, pretty_print=True)

                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'DOs':
                        fold = el1
                        for obj in UsoModule.select().where(UsoModule.module_type_id == '531'):
                            mtype = ModuleType.select().where(ModuleType.module_type_id == '531').get()
                            usoname = Uso.select().where(Uso.uso_id == obj.uso_id).get()
                            for el in fold.iter('{automation.control}object'):
                                if el.attrib['name'] == obj.uso_module_id:
                                    fold.remove(el)
                            name = obj.uso_module_id
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = name
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.modules.mod_DI"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.ModPosition"
                            atrb1.attrib['value'] = str(obj.sign)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.System.Attributes.Description"
                            atrb2.attrib['value'] = mtype.sign
                            object.append(atrb2)
                            atrb3 = etree.Element("attribute")
                            atrb3.attrib['type'] = "unit.Library.Attributes.ModID"
                            atrb3.attrib['value'] = mtype.hwdesc
                            object.append(atrb3)
                            atrb4 = etree.Element("attribute")
                            atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                            atrb4.attrib['value'] = usoname.description
                            object.append(atrb4)
                            fold.append(object)
                            tree.write(filename_omx, pretty_print=True)

                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'CPUs':
                        fold = el1
                        for obj in UsoModule.select().where(UsoModule.module_type_id == '502'):
                            mtype = ModuleType.select().where(ModuleType.module_type_id == '502').get()
                            usoname = Uso.select().where(Uso.uso_id == obj.uso_id).get()
                            for el in fold.iter('{automation.control}object'):
                                if el.attrib['name'] == obj.uso_module_id:
                                    fold.remove(el)
                            name = obj.uso_module_id
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = name
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.modules.mod_CPU"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.ModPosition"
                            atrb1.attrib['value'] = str(obj.sign)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.System.Attributes.Description"
                            atrb2.attrib['value'] = mtype.sign
                            object.append(atrb2)
                            atrb3 = etree.Element("attribute")
                            atrb3.attrib['type'] = "unit.Library.Attributes.ModID"
                            atrb3.attrib['value'] = mtype.hwdesc
                            object.append(atrb3)
                            atrb4 = etree.Element("attribute")
                            atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                            atrb4.attrib['value'] = usoname.description
                            object.append(atrb4)
                            fold.append(object)
                            tree.write(filename_omx, pretty_print=True)

                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'CPUs':
                        fold = el1
                        for obj in UsoModule.select().where(UsoModule.module_type_id == '503'):
                            mtype = ModuleType.select().where(ModuleType.module_type_id == '503').get()
                            usoname = Uso.select().where(Uso.uso_id == obj.uso_id).get()
                            for el in fold.iter('{automation.control}object'):
                                if el.attrib['name'] == obj.uso_module_id:
                                    fold.remove(el)
                            name = obj.uso_module_id
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = name
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.modules.mod_CPU"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.ModPosition"
                            atrb1.attrib['value'] = str(obj.sign)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.System.Attributes.Description"
                            atrb2.attrib['value'] = mtype.sign
                            object.append(atrb2)
                            atrb3 = etree.Element("attribute")
                            atrb3.attrib['type'] = "unit.Library.Attributes.ModID"
                            atrb3.attrib['value'] = mtype.hwdesc
                            object.append(atrb3)
                            atrb4 = etree.Element("attribute")
                            atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                            atrb4.attrib['value'] = usoname.description
                            object.append(atrb4)
                            fold.append(object)
                            tree.write(filename_omx, pretty_print=True)

                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'AIs':
                        fold = el1
                        for obj in UsoModule.select().where(UsoModule.module_type_id == '516'):
                            mtype = ModuleType.select().where(ModuleType.module_type_id == '516').get()
                            usoname = Uso.select().where(Uso.uso_id == obj.uso_id).get()
                            for el in fold.iter('{automation.control}object'):
                                if el.attrib['name'] == obj.uso_module_id:
                                    fold.remove(el)
                            name = obj.uso_module_id
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = name
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.modules.mod_AI"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.ModPosition"
                            atrb1.attrib['value'] = str(obj.sign)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.System.Attributes.Description"
                            atrb2.attrib['value'] = mtype.sign
                            object.append(atrb2)
                            atrb3 = etree.Element("attribute")
                            atrb3.attrib['type'] = "unit.Library.Attributes.ModID"
                            atrb3.attrib['value'] = mtype.hwdesc
                            object.append(atrb3)
                            atrb4 = etree.Element("attribute")
                            atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                            atrb4.attrib['value'] = usoname.description
                            object.append(atrb4)
                            fold.append(object)
                            tree.write(filename_omx, pretty_print=True)

                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'RSs':
                        fold = el1
                        for obj in UsoModule.select().where(UsoModule.module_type_id == '541'):
                            mtype = ModuleType.select().where(ModuleType.module_type_id == '541').get()
                            usoname = Uso.select().where(Uso.uso_id == obj.uso_id).get()
                            for el in fold.iter('{automation.control}object'):
                                if el.attrib['name'] == obj.uso_module_id:
                                    fold.remove(el)
                            name = obj.uso_module_id
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = name
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.modules.mod_RS"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.ModPosition"
                            atrb1.attrib['value'] = str(obj.sign)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.System.Attributes.Description"
                            atrb2.attrib['value'] = mtype.sign
                            object.append(atrb2)
                            atrb3 = etree.Element("attribute")
                            atrb3.attrib['type'] = "unit.Library.Attributes.ModID"
                            atrb3.attrib['value'] = mtype.hwdesc
                            object.append(atrb3)
                            atrb4 = etree.Element("attribute")
                            atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                            atrb4.attrib['value'] = usoname.description
                            object.append(atrb4)
                            fold.append(object)
                            tree.write(filename_omx, pretty_print=True)

                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'CNs':
                        fold = el1
                        for obj in UsoModule.select().where(UsoModule.module_type_id == '545'):
                            mtype = ModuleType.select().where(ModuleType.module_type_id == '545').get()
                            usoname = Uso.select().where(Uso.uso_id == obj.uso_id).get()
                            for el in fold.iter('{automation.control}object'):
                                if el.attrib['name'] == obj.uso_module_id:
                                    fold.remove(el)
                            name = obj.uso_module_id
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = name
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.modules.mod_CN"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.ModPosition"
                            atrb1.attrib['value'] = str(obj.sign)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.System.Attributes.Description"
                            atrb2.attrib['value'] = mtype.sign
                            object.append(atrb2)
                            atrb3 = etree.Element("attribute")
                            atrb3.attrib['type'] = "unit.Library.Attributes.ModID"
                            atrb3.attrib['value'] = mtype.hwdesc
                            object.append(atrb3)
                            atrb4 = etree.Element("attribute")
                            atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                            atrb4.attrib['value'] = usoname.description
                            object.append(atrb4)
                            fold.append(object)
                            tree.write(filename_omx, pretty_print=True)

                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'MNs':
                        fold = el1
                        for obj in UsoModule.select().where(UsoModule.module_type_id == '546'):
                            mtype = ModuleType.select().where(ModuleType.module_type_id == '546').get()
                            usoname = Uso.select().where(Uso.uso_id == obj.uso_id).get()
                            for el in fold.iter('{automation.control}object'):
                                if el.attrib['name'] == obj.uso_module_id:
                                    fold.remove(el)
                            name = obj.uso_module_id
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = name
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.modules.mod_CN"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.ModPosition"
                            atrb1.attrib['value'] = str(obj.sign)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.System.Attributes.Description"
                            atrb2.attrib['value'] = mtype.sign
                            object.append(atrb2)
                            atrb3 = etree.Element("attribute")
                            atrb3.attrib['type'] = "unit.Library.Attributes.ModID"
                            atrb3.attrib['value'] = mtype.hwdesc
                            object.append(atrb3)
                            atrb4 = etree.Element("attribute")
                            atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                            atrb4.attrib['value'] = usoname.description
                            object.append(atrb4)
                            fold.append(object)
                            tree.write(filename_omx, pretty_print=True)

                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'PSUs':
                        fold = el1
                        for obj in UsoModule.select().where(UsoModule.module_type_id == '550'):
                            mtype = ModuleType.select().where(ModuleType.module_type_id == '550').get()

                            usoname = Uso.select().where(Uso.uso_id == obj.uso_id).get()
                            for el in fold.iter('{automation.control}object'):
                                if el.attrib['name'] == obj.uso_module_id:
                                    fold.remove(el)
                            name = obj.uso_module_id
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = name
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.modules.mod_PSU"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.ModPosition"
                            atrb1.attrib['value'] = str(obj.sign)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.System.Attributes.Description"
                            atrb2.attrib['value'] = mtype.sign
                            object.append(atrb2)
                            atrb3 = etree.Element("attribute")
                            atrb3.attrib['type'] = "unit.Library.Attributes.ModID"
                            atrb3.attrib['value'] = mtype.hwdesc
                            object.append(atrb3)
                            atrb4 = etree.Element("attribute")
                            atrb4.attrib['type'] = "unit.Library.Attributes.ModUSO"
                            atrb4.attrib['value'] = usoname.description
                            object.append(atrb4)
                            fold.append(object)
                            tree.write(filename_omx, pretty_print=True)
    except:
        print('не добавлено')

def add_MapTagName(filename_omx):


    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(filename_omx, parser)
    root = tree.getroot()
    shutil.copy(filename_omx, filename_omx + '_backup')
    try:
        for obj in Kzfkp.select():
            if obj.hwdesc in ['521','531']:

                type = modId[obj.hwdesc]
                for elem in root.findall('item'):
                    ch = str(obj.channel)
                    if str(obj.channel) in channel:
                        ch = '0'+ str(obj.channel)
                    strid = "Root.Diag." + type +"s." +  obj.signal_id[:-3] + ".ch_DI_" +ch
                    if elem.attrib['id'] == strid:
                        root.remove(elem)
                apl = root
                object = etree.Element('item')
                object.attrib['id'] = strid
                object.attrib['value'] = str(obj.tag)
                apl.append(object)

        for obj in Kzfkp.select():
            if obj.hwdesc in ['514','516']:

                type = modId[obj.hwdesc]
                for elem in root.findall('item'):
                    ch = str(obj.channel)
                    if str(obj.channel) in channel:
                        ch = '0'+ str(obj.channel)
                    strid = "Root.Diag." + type +"s." +  obj.signal_id[:-3] + ".ch_AI_" +ch
                    if elem.attrib['id'] == strid:
                        root.remove(elem)
                apl = root
                object = etree.Element('item')
                object.attrib['id'] = strid
                object.attrib['value'] = str(obj.tag)
                apl.append(object)


        tree.write(filename_omx, pretty_print=True)
        print('Создан')

    except:
        print('какой то косяк')

def add_MapSignalName(filename_omx):


    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(filename_omx, parser)
    root = tree.getroot()
    shutil.copy(filename_omx, filename_omx + '_backup')
    try:
        for obj in Kzfkp.select():
            if obj.hwdesc in ['521','531']:

                type = modId[obj.hwdesc]
                print(type)
                ch = str(obj.channel)
                if str(obj.channel) in channel:
                    ch = '0' + str(obj.channel)
                strid = "Root.Diag." + type + "s." + obj.signal_id[:-3] + ".ch_DI_" + ch
                for elem in root.findall('item'):


                    if elem.attrib['id'] == strid:
                        root.remove(elem)
                apl = root
                object = etree.Element('item')
                object.attrib['id'] = strid
                object.attrib['value'] = str(obj.signalname)
                apl.append(object)

        for obj in Kzfkp.select():
            if obj.hwdesc in ['514','516']:

                type = modId[obj.hwdesc]
                for elem in root.findall('item'):
                    ch = str(obj.channel)
                    if str(obj.channel) in channel:
                        ch = '0'+ str(obj.channel)
                    strid = "Root.Diag." + type +"s." +  obj.signal_id[:-3] + ".ch_AI_" +ch
                    if elem.attrib['id'] == strid:
                        root.remove(elem)
                apl = root
                object = etree.Element('item')
                object.attrib['id'] = strid
                object.attrib['value'] = str(obj.signalname)
                apl.append(object)


        tree.write(filename_omx, pretty_print=True)
        print('Создан')

    except:
        print('какой то косяк')

def add_Mapklk(filename_omx):


    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(filename_omx, parser)
    root = tree.getroot()
    shutil.copy(filename_omx, filename_omx + '_backup')
    try:
        for obj in Kzfkp.select():
            if obj.hwdesc in ['521','531']:

                type = modId[obj.hwdesc]
                print(type)
                ch = str(obj.channel)
                if str(obj.channel) in channel:
                    ch = '0' + str(obj.channel)
                strid = "Root.Diag." + type + "s." + obj.signal_id[:-3] + ".ch_DI_" + ch
                for elem in root.findall('item'):


                    if elem.attrib['id'] == strid:
                        root.remove(elem)
                apl = root
                object = etree.Element('item')
                object.attrib['id'] = strid
                object.attrib['value'] = str(obj.terminalblock)
                apl.append(object)

        for obj in Kzfkp.select():
            if obj.hwdesc in ['514','516']:

                type = modId[obj.hwdesc]
                for elem in root.findall('item'):
                    ch = str(obj.channel)
                    if str(obj.channel) in channel:
                        ch = '0'+ str(obj.channel)
                    strid = "Root.Diag." + type +"s." +  obj.signal_id[:-3] + ".ch_AI_" +ch
                    if elem.attrib['id'] == strid:
                        root.remove(elem)
                apl = root
                object = etree.Element('item')
                object.attrib['id'] = strid
                object.attrib['value'] = str(obj.terminalblock)
                apl.append(object)


        tree.write(filename_omx, pretty_print=True)
        print('Создан')

    except:
        print('какой то косяк')

def add_Mapkont(filename_omx):


    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(filename_omx, parser)
    root = tree.getroot()
    shutil.copy(filename_omx, filename_omx + '_backup')
    try:
        for obj in Kzfkp.select():
            if obj.hwdesc in ['521','531']:

                type = modId[obj.hwdesc]
                print(type)
                ch = str(obj.channel)
                if str(obj.channel) in channel:
                    ch = '0' + str(obj.channel)
                strid = "Root.Diag." + type + "s." + obj.signal_id[:-3] + ".ch_DI_" + ch
                for elem in root.findall('item'):


                    if elem.attrib['id'] == strid:
                        root.remove(elem)
                apl = root
                object = etree.Element('item')
                object.attrib['id'] = strid
                object.attrib['value'] = str(obj.kont)
                apl.append(object)

        for obj in Kzfkp.select():
            if obj.hwdesc in ['514','516']:

                type = modId[obj.hwdesc]
                for elem in root.findall('item'):
                    ch = str(obj.channel)
                    if str(obj.channel) in channel:
                        ch = '0'+ str(obj.channel)
                    strid = "Root.Diag." + type +"s." +  obj.signal_id[:-3] + ".ch_AI_" +ch
                    if elem.attrib['id'] == strid:
                        root.remove(elem)
                apl = root
                object = etree.Element('item')
                object.attrib['id'] = strid
                object.attrib['value'] = str(obj.kont)
                apl.append(object)


        tree.write(filename_omx, pretty_print=True)
        print('Создан')

    except:
        print('какой то косяк')


def Read_Oip():
    try:
        Analog.create_table()
        wb = openpyxl.load_workbook("E:\sandbox\MK500\oip.xlsx")
        for sheet in wb.worksheets:
            if sheet.title == 'oip':
                data = []
                a_dict = []
                rows = sheet.max_row
                print(rows)
                for i in range(2, rows + 1):
                    a_dict = dict(  tag = sheet.cell(row=i, column=5).value,
                                    description = sheet.cell(row=i, column=4).value,
                                    egu = sheet.cell(row=i, column=9).value,
                                    sign = sheet.cell(row=i, column=7).value,
                                    index_arr = sheet.cell(row=i, column=2).value,
                                    zone = sheet.cell(row=i, column=3).value,
                                    msggrp = sheet.cell(row=i, column=8).value)

                    data.append(a_dict)

        base_write_many2(Analog, data)
    except:
        print("не удалось")

def Read_Msggrp():
    try:
        Ustgrp.create_table()
        wb = openpyxl.load_workbook("E:\sandbox\MK500\oip.xlsx")
        for sheet in wb.worksheets:
            if sheet.title == 'ust':
                data = []
                a_dict = []
                rows = sheet.max_row
                print(rows)
                for i in range(2, rows + 1):
                    a_dict = dict(      msggrp_id = sheet.cell(row=i, column=1).value,
                                        name = sheet.cell(row=i, column=2).value,
                                        min6 = sheet.cell(row=i, column=3).value,
                                        min5 = sheet.cell(row=i, column=4).value,
                                        min4 = sheet.cell(row=i, column=5).value,
                                        min3 = sheet.cell(row=i, column=6).value,
                                        min2 = sheet.cell(row=i, column=7).value,
                                        min1 = sheet.cell(row=i, column=8).value,
                                        max1 = sheet.cell(row=i, column=9).value,
                                        max2 = sheet.cell(row=i, column=10).value,
                                        max3 = sheet.cell(row=i, column=11).value,
                                        max4 = sheet.cell(row=i, column=12).value,
                                        max5 = sheet.cell(row=i, column=13).value,
                                        max6 = sheet.cell(row=i, column=14).value)
                    data.append(a_dict)
        base_write_many2(Ustgrp, data)
    except:
        print("не удалось")

def add_oip(filename_omx):
    try:
        print(filename_omx)
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(filename_omx, parser)
        root = tree.getroot()
        for el in root.iter('{automation.deployment}application-object'):
            if el.attrib['name'] == "Application_PLC":
                apl = el
                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'Analogs':
                        fold = el1

                        for obj in Analog.select():
                            print(obj.tag)
                            for el in fold.iter('{automation.control}object'):
                                    if el.attrib['name'] == obj.tag:
                                        print(obj.tag)
                                        fold.remove(el)
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = obj.tag
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.Analog_PLC"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.Index"
                            atrb1.attrib['value'] = str(obj.index_arr)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.Library.Attributes.Sign"
                            atrb2.attrib['value'] = obj.sign
                            object.append(atrb2)
                            atrb3 = etree.Element("attribute")
                            atrb3.attrib['type'] = "unit.Library.Attributes.EGU_Desc"
                            atrb3.attrib['value'] = str(obj.egu)
                            object.append(atrb3)
                            atrb4 = etree.Element("attribute")
                            atrb4.attrib['type'] = "unit.Library.Attributes.EGU_Desc_phys"
                            atrb4.attrib['value'] = "мкА"
                            object.append(atrb4)
                            atrb5 = etree.Element("attribute")
                            atrb5.attrib['type'] = "unit.System.Attributes.Description"
                            atrb5.attrib['value'] = obj.description
                            object.append(atrb5)
                            fold.append(object)
                            tree.write(filename_omx,pretty_print=True)


    except:
        print('не добавлено')

def Read_Aux():
    try:
        AuxSystems.create_table()
        wb = openpyxl.load_workbook("E:\sandbox\MK500\oip.xlsx")
        for sheet in wb.worksheets:
            if sheet.title == 'Aux':
                data = []
                a_dict = []
                rows = sheet.max_row
                print(rows)
                for i in range(2, rows + 1):
                    a_dict = dict(          tag = sheet.cell(row=i, column=5).value,
                                            description = sheet.cell(row=i, column=4).value,
                                            sign = sheet.cell(row=i, column=7).value,
                                            index_arr = sheet.cell(row=i, column=2).value,
                                            zone = sheet.cell(row=i, column=3).value)
                    data.append(a_dict)
        base_write_many2(AuxSystems, data)
    except:
        print("не удалось")

def add_auxsystem(filename_omx):
    try:
        print(filename_omx)
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(filename_omx, parser)
        root = tree.getroot()
        for el in root.iter('{automation.deployment}application-object'):
            if el.attrib['name'] == "Application_PLC":
                apl = el
                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'AuxSystems':
                        fold = el1

                        for obj in AuxSystems.select():
                            print(obj.tag)
                            for el in fold.iter('{automation.control}object'):
                                    if el.attrib['name'] == obj.tag:
                                        print(obj.tag)
                                        fold.remove(el)
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = obj.tag
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.AuxSystem_PLC"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.Index"
                            atrb1.attrib['value'] = str(obj.index_arr)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.Library.Attributes.Sign"
                            atrb2.attrib['value'] = obj.sign
                            object.append(atrb2)
                            atrb5 = etree.Element("attribute")
                            atrb5.attrib['type'] = "unit.System.Attributes.Description"
                            atrb5.attrib['value'] = obj.description
                            object.append(atrb5)
                            fold.append(object)
                            tree.write(filename_omx,pretty_print=True)


    except:
        print('не добавлено')

def Read_Discret():
    try:
        Discret.create_table()
        wb = openpyxl.load_workbook("E:\sandbox\MK500\oip.xlsx")
        for sheet in wb.worksheets:
            if sheet.title == 'dsc':
                data = []
                a_dict = []
                rows = sheet.max_row
                print(rows)
                for i in range(2, rows + 1):
                    a_dict = dict(          tag = sheet.cell(row=i, column=5).value,
                                            description = sheet.cell(row=i, column=4).value,
                                            sign = sheet.cell(row=i, column=6).value,
                                            index_arr = sheet.cell(row=i, column=2).value,
                                            zone = sheet.cell(row=i, column=3).value,
                                            colorsheme = "2")
                    data.append(a_dict)
        base_write_many2(Discret, data)
    except:
        print("не удалось")

def add_discret(filename_omx):
    try:
        print(filename_omx)
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(filename_omx, parser)
        root = tree.getroot()
        for el in root.iter('{automation.deployment}application-object'):
            if el.attrib['name'] == "Application_PLC":
                apl = el
                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'Diskrets':
                        fold = el1

                        for obj in Discret.select():
                            print(obj.tag)
                            for el in fold.iter('{automation.control}object'):
                                    if el.attrib['name'] == obj.tag:
                                        print(obj.tag)
                                        fold.remove(el)
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = str(obj.tag)
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.Diskret_PLC"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.Index"
                            atrb1.attrib['value'] = str(obj.index_arr)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.Library.Attributes.Sign"
                            atrb2.attrib['value'] = str(obj.sign)
                            object.append(atrb2)
                            atrb5 = etree.Element("attribute")
                            atrb5.attrib['type'] = "unit.System.Attributes.Description"
                            atrb5.attrib['value'] = str(obj.description)
                            object.append(atrb5)
                            fold.append(object)
                            tree.write(filename_omx,pretty_print=True)


    except:
        print('не добавлено')

def Read_UTS():
    try:
        UTS.create_table()
        wb = openpyxl.load_workbook("E:\sandbox\MK500\oip.xlsx")
        for sheet in wb.worksheets:
            if sheet.title == 'uts':
                data = []
                a_dict = []
                rows = sheet.max_row
                print(rows)
                for i in range(2, rows + 1):
                    a_dict = dict(          tag = sheet.cell(row=i, column=5).value,
                                            description = sheet.cell(row=i, column=4).value,
                                            sign = sheet.cell(row=i, column=7).value,
                                            index_arr = sheet.cell(row=i, column=2).value,
                                            zone = sheet.cell(row=i, column=3).value)
                    data.append(a_dict)
        base_write_many2(UTS, data)
    except:
        print("не удалось")

def add_uts(filename_omx):
    try:
        print(filename_omx)
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(filename_omx, parser)
        root = tree.getroot()
        for el in root.iter('{automation.deployment}application-object'):
            if el.attrib['name'] == "Application_PLC":
                apl = el
                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'UTSs':
                        fold = el1

                        for obj in UTS.select():
                            print(obj.tag)
                            for el in fold.iter('{automation.control}object'):
                                    if el.attrib['name'] == obj.tag:
                                        print(obj.tag)
                                        fold.remove(el)
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = str(obj.tag)
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.UTS_PLC"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.Index"
                            atrb1.attrib['value'] = str(obj.index_arr)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.Library.Attributes.Sign"
                            atrb2.attrib['value'] = str(obj.sign)
                            object.append(atrb2)
                            atrb5 = etree.Element("attribute")
                            atrb5.attrib['type'] = "unit.System.Attributes.Description"
                            atrb5.attrib['value'] = str(obj.description)
                            object.append(atrb5)
                            fold.append(object)
                            tree.write(filename_omx,pretty_print=True)


    except:
        print('не добавлено')

def Read_Valves():
    try:
        Valves.create_table()
        wb = openpyxl.load_workbook("E:\sandbox\MK500\oip.xlsx")
        for sheet in wb.worksheets:
            if sheet.title == 'valves':
                data = []
                a_dict = []
                rows = sheet.max_row
                print(rows)
                for i in range(2, rows + 1):
                    a_dict = dict(          tag = sheet.cell(row=i, column=5).value,
                                            description = sheet.cell(row=i, column=4).value,
                                            sign = sheet.cell(row=i, column=6).value,
                                            index_arr = sheet.cell(row=i, column=2).value,
                                            zone = sheet.cell(row=i, column=3).value)
                    data.append(a_dict)
        base_write_many2(Valves, data)
    except:
        print("не удалось")

def add_valves(filename_omx):
    try:
        print(filename_omx)
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(filename_omx, parser)
        root = tree.getroot()
        for el in root.iter('{automation.deployment}application-object'):
            if el.attrib['name'] == "Application_PLC":
                apl = el
                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'Valves':
                        fold = el1

                        for obj in Valves.select():
                            print(obj.tag)
                            for el in fold.iter('{automation.control}object'):
                                    if el.attrib['name'] == obj.tag:
                                        print(obj.tag)
                                        fold.remove(el)
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = str(obj.tag)
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.Valve_PLC"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.Index"
                            atrb1.attrib['value'] = str(obj.index_arr)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.Library.Attributes.Sign"
                            atrb2.attrib['value'] = str(obj.sign)
                            object.append(atrb2)
                            atrb5 = etree.Element("attribute")
                            atrb5.attrib['type'] = "unit.System.Attributes.Description"
                            atrb5.attrib['value'] = str(obj.description)
                            object.append(atrb5)
                            fold.append(object)
                            tree.write(filename_omx,pretty_print=True)


    except:
        print('не добавлено')

def Read_NA():
    try:
        NA.create_table()
        wb = openpyxl.load_workbook("E:\sandbox\MK500\oip.xlsx")
        for sheet in wb.worksheets:
            if sheet.title == 'NA':
                data = []
                a_dict = []
                rows = sheet.max_row
                print(rows)
                for i in range(2, rows + 1):
                    a_dict = dict(          tag = sheet.cell(row=i, column=5).value,
                                            description = sheet.cell(row=i, column=4).value,
                                            sign = sheet.cell(row=i, column=6).value,
                                            index_arr = sheet.cell(row=i, column=2).value,
                                            zone = sheet.cell(row=i, column=3).value)
                    data.append(a_dict)
        base_write_many2(NA, data)
    except:
        print("не удалось")

def add_na(filename_omx):
    try:
        print(filename_omx)
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(filename_omx, parser)
        root = tree.getroot()
        for el in root.iter('{automation.deployment}application-object'):
            if el.attrib['name'] == "Application_PLC":
                apl = el
                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'NAs':
                        fold = el1

                        for obj in NA.select():
                            print(obj.tag)
                            for el in fold.iter('{automation.control}object'):
                                    if el.attrib['name'] == obj.tag:
                                        print(obj.tag)
                                        fold.remove(el)
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = str(obj.tag)
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.NA_PLC"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb1 = etree.Element("attribute")
                            atrb1.attrib['type'] = "unit.Library.Attributes.Index"
                            atrb1.attrib['value'] = str(obj.index_arr)
                            object.append(atrb1)
                            atrb2 = etree.Element("attribute")
                            atrb2.attrib['type'] = "unit.Library.Attributes.Sign"
                            atrb2.attrib['value'] = str(obj.sign)
                            object.append(atrb2)
                            atrb5 = etree.Element("attribute")
                            atrb5.attrib['type'] = "unit.System.Attributes.Description"
                            atrb5.attrib['value'] = str(obj.description)
                            object.append(atrb5)
                            fold.append(object)
                            tree.write(filename_omx,pretty_print=True)


    except:
        print('не добавлено')

def add_ktpr(filename_omx):
    try:
        print(filename_omx)
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(filename_omx, parser)
        root = tree.getroot()
        for el in root.iter('{automation.deployment}application-object'):
            if el.attrib['name'] == "Application_PLC":
                apl = el
                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'KTPRs':
                        fold = el1
                        i=0
                        for i in range(1,25):
                            for el in fold.iter('{automation.control}object'):
                                if el.attrib['name'] == "Group_" + str(i):
                                    print("Group_"+str(i))
                                    fold.remove(el)
                            object = etree.Element("{automation.control}object")
                            object.attrib['name'] = "Group_" + str(i)
                            object.attrib['uuid'] = str(uuid.uuid1())
                            object.attrib['base-type'] = "unit.Library.PLC_Types.KTPRx_PLC"
                            object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                            atrb5 = etree.Element("attribute")
                            atrb5.attrib['type'] = "unit.System.Attributes.Description"
                            atrb5.attrib['value'] = "Регистр состояния защит (4 защиты)" + "группа " + str(i)
                            object.append(atrb5)
                            fold.append(object)
                            tree.write(filename_omx,pretty_print=True)


    except:
        print('не добавлено')

def add_ktpras(filename_omx):
    try:
        print(filename_omx)
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(filename_omx, parser)
        root = tree.getroot()
        for el in root.iter('{automation.deployment}application-object'):
            if el.attrib['name'] == "Application_PLC":
                apl = el
                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'KTPRAs':
                        fold = el1
                        for j in range(1,5):
                            for na_fold in fold.iter('{automation.control}object'):
                                if na_fold.attrib['name']=="NA_"+str(j):
                                    for i in range(1,25):
                                        for el in na_fold.iter('{automation.control}object'):
                                            if el.attrib['name'] == "Group_" + str(i):
                                                na_fold.remove(el)
                                        object = etree.Element("{automation.control}object")
                                        object.attrib['name'] = "Group_" + str(i)
                                        object.attrib['uuid'] = str(uuid.uuid1())
                                        object.attrib['base-type'] = "unit.Library.PLC_Types.KTPRx_PLC"
                                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                                        atrb5 = etree.Element("attribute")
                                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                                        atrb5.attrib['value'] = "Регистр состояния защит (4 защиты)" + "группа " + str(i)
                                        object.append(atrb5)
                                        na_fold.append(object)
                                        tree.write(filename_omx,pretty_print=True)


    except:
        print('не добавлено')

def add_gmpnas(filename_omx):
    try:
        print(filename_omx)
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(filename_omx, parser)
        root = tree.getroot()
        for el in root.iter('{automation.deployment}application-object'):
            if el.attrib['name'] == "Application_PLC":
                apl = el
                for el1 in apl.iter('{automation.control}object'):
                    if el1.attrib['name'] == 'GMPNAs':
                        fold = el1
                        for j in range(1,5):
                            for na_fold in fold.iter('{automation.control}object'):
                                if na_fold.attrib['name']=="NA_"+str(j):
                                    for i in range(1,25):
                                        for el in na_fold.iter('{automation.control}object'):
                                            if el.attrib['name'] == "Group_" + str(i):
                                                na_fold.remove(el)
                                        object = etree.Element("{automation.control}object")
                                        object.attrib['name'] = "Group_" + str(i)
                                        object.attrib['uuid'] = str(uuid.uuid1())
                                        object.attrib['base-type'] = "unit.Library.PLC_Types.GMPNA_PLC"
                                        object.attrib['aspect'] = "unit.Library.PLC_Types.PLC"
                                        atrb5 = etree.Element("attribute")
                                        atrb5.attrib['type'] = "unit.System.Attributes.Description"
                                        atrb5.attrib['value'] = "Регистр состояния готовностей (4 готовности)" + "группа " + str(i)
                                        object.append(atrb5)
                                        na_fold.append(object)
                                        tree.write(filename_omx,pretty_print=True)


    except:
        print('не добавлено')



def add_map_opcua(filename_ua):


    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(filename_ua, parser)
    root = tree.getroot()
    i = 0
    map_zd = {'StateValve':'stateZD','StateValveEx':'stateZDExt','ConfigValve':'ConfigZD','Tm.tmZD':'tmZD'}
    map_mna = {'StateNA':'stateNA',
                'operatingTimeSinceSwitchingOn': 'StatNA_OpTimeSwOn',
                'operatingTimeSinceSwitchingOnSet': 'StatNA_OpTimeSwOnSet',
                'operatingTimeBeforeOverhaul': 'StatNA_OpTimeBeforeOvhl',
                'operatingTimeBeforeOverhaulSet': 'StatNA_OpTimeBeforeOvhlSet',
                'numOfStart': 'StatNA_NumOfStarts',
                'dateTimeOfStart': 'StatNA_DateTimeOfStart',
                'dateTimeOfStop': 'StatNA_DateTimeOfStop',
                'operatingTimeCurrentMonth': 'StatNA_OpTimeCurrMonth',
                'operatingTimeLastMonth': 'StatNA_OpTimeLastMonth'}
    map_dio = {'mod_State': 'dDioState',

                'DError': 'dError',
                'DData': 'dData'}
    map_ch = {'ch_01': 'mA',
                'ch_02': 'dDioADatamA',
                'ch_03': 'dDioADatamA',
                'ch_04': 'dDioADatamA',
                'ch_05': 'dDioADatamA',
                'ch_06': 'dDioADatamA',
                'ch_07': 'dDioADatamA',
                'ch_08': 'dDioADatamA'}
    try:
        for obj in Analog.select():
                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.AIVisualValue'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'AIVisualValue['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.UstEnable'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'UstEnable['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.RangeBottom'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'RangeBottom['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.ScaleBottom'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'ScaleBottom['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.ScaleExtBottom'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'ScaleExtBottom['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.RangeTop'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'RangeTop['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.ScaleTop'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'ScaleTop['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.ScaleExtTop'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'ScaleExtTop['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.T'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'T['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.Hist'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'Hist['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.UstMin6'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'UstMin6['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.UstMin5'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'UstMin5['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.UstMin4'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'UstMin4['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.UstMin3'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'UstMin3['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.UstMin2'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'UstMin2['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.UstMin1'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'UstMin1['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.UstMax6'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'UstMax6['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.UstMax5'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'UstMax5['+str(obj.index_arr) +']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.UstMax4'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'UstMax4[' + str(obj.index_arr) + ']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.UstMax3'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'UstMax3[' + str(obj.index_arr) + ']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.UstMax2'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'UstMax2[' + str(obj.index_arr) + ']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.Presets_PLC.UstMax1'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'UstMax1[' + str(obj.index_arr) + ']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.AIElValue'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'AIElValue[' + str(obj.index_arr) + ']'
                object.append(nodeId)
                apl.append(object)

                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.AIValue'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'AIValue[' + str(obj.index_arr) + ']'
                object.append(nodeId)
                apl.append(object)


                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Analogs.' + str(obj.tag) + '.StateAI'
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = 'StateAI[' + str(obj.index_arr) + ']'
                object.append(nodeId)
                apl.append(object)

        for obj in Discret.select():
            apl = root
            object = etree.Element('item')
            object.attrib['Binding'] = 'Introduced'
            np = etree.Element('node-path')
            np.text = 'Root.Diskrets.' + str(obj.tag) + '.StateDI'
            object.append(np)
            namespace = etree.Element('namespace')
            namespace.text = "nft"
            object.append(namespace)
            nodeType = etree.Element('nodeIdType')
            nodeType.text = 'String'
            object.append(nodeType)
            nodeId = etree.Element('nodeId')
            nodeId.text = 'StateDI[' + str(obj.index_arr) + ']'
            object.append(nodeId)
            apl.append(object)


        for zd in map_zd:
            for obj in Valves.select():
                ndId = map_zd[zd]+'['+ str(obj.index_arr) + ']'
                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Valves.' + str(obj.tag) + '.' + str(zd)
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = ndId
                object.append(nodeId)
                apl.append(object)

        for mna in map_mna:
            for obj in NA.select():
                ndId = map_mna[mna]+'['+ str(obj.index_arr) + ']'
                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.NAs.' + str(obj.tag) + '.' + str(mna)
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = ndId
                object.append(nodeId)
                apl.append(object)

        for dio in map_dio:
            for obj in UsoModule.select():
                ndId = map_dio[dio]+'['+ str(obj.index_arr) + ']'
                apl = root
                object = etree.Element('item')
                object.attrib['Binding'] = 'Introduced'
                np = etree.Element('node-path')
                np.text = 'Root.Diag.' + modId[obj.module_type_id] + 's.' + str(obj.uso_module_id) + '.' + str(dio)
                object.append(np)
                namespace = etree.Element('namespace')
                namespace.text = "nft"
                object.append(namespace)
                nodeType = etree.Element('nodeIdType')
                nodeType.text = 'String'
                object.append(nodeType)
                nodeId = etree.Element('nodeId')
                nodeId.text = ndId
                object.append(nodeId)
                apl.append(object)

        j = 0
        for obj in UsoModule.select():

            i = 0
            if obj.module_type_id in ['514','516']:
                j=j+1
                for dio_ch in map_ch:
                    i=i+1
                    ndId = map_ch[dio_ch]+'['+ str((j-1)*8+i) + ']'
                    apl = root
                    object = etree.Element('item')
                    object.attrib['Binding'] = 'Introduced'
                    np = etree.Element('node-path')
                    np.text = 'Root.Diag.' + modId[obj.module_type_id] + 's.' + str(obj.uso_module_id) + '.' + dio_ch
                    object.append(np)
                    namespace = etree.Element('namespace')
                    namespace.text = "nft"
                    object.append(namespace)
                    nodeType = etree.Element('nodeIdType')
                    nodeType.text = 'String'
                    object.append(nodeType)
                    nodeId = etree.Element('nodeId')
                    nodeId.text = ndId
                    object.append(nodeId)
                    apl.append(object)
                    #print('Root.Diag.' + modId[obj.module_type_id] + 's.' + str(obj.uso_module_id) + '.' + dio_ch)






        tree.write(filename_ua, pretty_print=True)

        print('Создан')
    except:
        print('какой то косяк')

def add_attrib_signal_name(filename):
    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(filename, parser)
    root = tree.getroot()

    for obj in Kzfkp.select().where(Kzfkp.typesignal == 'AI'):
        apl = root
        object = etree.Element('item')
        object.attrib['id'] = 'Root.Diag.AIs.' + str(obj.signal_id)[:-3] + '.ch_AI_0'+str(obj.channel)
        object.attrib['value'] = str(obj.signalname)
        apl.append(object)
        tree.write(filename, pretty_print=True)

    for obj in Kzfkp.select().where(Kzfkp.typesignal == 'DI'):
        nchannel = str(obj.channel)
        if obj.channel<10:
            nchannel = '0'+str(obj.channel)
        apl = root
        object = etree.Element('item')
        object.attrib['id'] = 'Root.Diag.DIs.' + str(obj.signal_id)[:-3] + '.ch_DI_'+nchannel
        object.attrib['value'] = str(obj.signalname)
        apl.append(object)
        tree.write(filename, pretty_print=True)

    for obj in Kzfkp.select().where(Kzfkp.typesignal == 'DO'):
        nchannel = str(obj.channel)
        if obj.channel<10:
            nchannel = '0'+str(obj.channel)
        apl = root
        object = etree.Element('item')
        object.attrib['id'] = 'Root.Diag.DOs.' + str(obj.signal_id)[:-3] + '.ch_DI_'+nchannel
        object.attrib['value'] = str(obj.signalname)
        apl.append(object)
        tree.write(filename, pretty_print=True)

def add_attrib_klk(filename):
    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(filename, parser)
    root = tree.getroot()

    for obj in Kzfkp.select().where(Kzfkp.typesignal == 'AI'):
        apl = root
        val = str(obj.terminalblock)
        if val in {'None','none'}:
            val = ' '
        object = etree.Element('item')
        object.attrib['id'] = 'Root.Diag.AIs.' + str(obj.signal_id)[:-3] + '.ch_AI_0'+str(obj.channel)
        object.attrib['value'] = val
        apl.append(object)
        tree.write(filename, pretty_print=True)

    for obj in Kzfkp.select().where(Kzfkp.typesignal == 'DI'):
        nchannel = str(obj.channel)
        val = str(obj.terminalblock)
        if val in {'None','none'}:
            val = ' '
        if obj.channel<10:
            nchannel = '0'+str(obj.channel)
        apl = root
        object = etree.Element('item')
        object.attrib['id'] = 'Root.Diag.DIs.' + str(obj.signal_id)[:-3] + '.ch_DI_'+nchannel
        object.attrib['value'] = val
        apl.append(object)
        tree.write(filename, pretty_print=True)

    for obj in Kzfkp.select().where(Kzfkp.typesignal == 'DO'):
        nchannel = str(obj.channel)
        val = str(obj.terminalblock)
        if val in {'None','none'}:
            val = ' '
        if obj.channel<10:
            nchannel = '0'+str(obj.channel)
        apl = root
        object = etree.Element('item')
        object.attrib['id'] = 'Root.Diag.DOs.' + str(obj.signal_id)[:-3] + '.ch_DI_'+nchannel
        object.attrib['value'] = val
        apl.append(object)
        tree.write(filename, pretty_print=True)

def add_attrib_kont(filename):
    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(filename, parser)
    root = tree.getroot()

    for obj in Kzfkp.select().where(Kzfkp.typesignal == 'AI'):
        apl = root
        val = str(obj.kont)
        if val in {'None','none'}:
            val = ' '
        object = etree.Element('item')
        object.attrib['id'] = 'Root.Diag.AIs.' + str(obj.signal_id)[:-3] + '.ch_AI_0'+str(obj.channel)
        object.attrib['value'] = val
        apl.append(object)
        tree.write(filename, pretty_print=True)

    for obj in Kzfkp.select().where(Kzfkp.typesignal == 'DI'):
        nchannel = str(obj.channel)
        val = str(obj.kont)
        if val in {'None','none'}:
            val = ' '
        if obj.channel<10:
            nchannel = '0'+str(obj.channel)
        apl = root
        object = etree.Element('item')
        object.attrib['id'] = 'Root.Diag.DIs.' + str(obj.signal_id)[:-3] + '.ch_DI_'+nchannel
        object.attrib['value'] = val
        apl.append(object)
        tree.write(filename, pretty_print=True)

    for obj in Kzfkp.select().where(Kzfkp.typesignal == 'DO'):
        nchannel = str(obj.channel)
        val = str(obj.kont)
        if val in {'None','none'}:
            val = ' '
        if obj.channel<10:
            nchannel = '0'+str(obj.channel)
        apl = root
        object = etree.Element('item')
        object.attrib['id'] = 'Root.Diag.DOs.' + str(obj.signal_id)[:-3] + '.ch_DI_'+nchannel
        object.attrib['value'] = val
        apl.append(object)
        tree.write(filename, pretty_print=True)

def add_attrib_tagname(filename):
    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(filename, parser)
    root = tree.getroot()

    for obj in Kzfkp.select().where(Kzfkp.typesignal == 'AI'):
        apl = root
        val = str(obj.tag)
        if val in {'None','none'}:
            val = ' '
        object = etree.Element('item')
        object.attrib['id'] = 'Root.Diag.AIs.' + str(obj.signal_id)[:-3] + '.ch_AI_0'+str(obj.channel)
        object.attrib['value'] = val
        apl.append(object)
        tree.write(filename, pretty_print=True)

    for obj in Kzfkp.select().where(Kzfkp.typesignal == 'DI'):
        nchannel = str(obj.channel)
        val = str(obj.tag)
        if val in {'None','none'}:
            val = ' '
        if obj.channel<10:
            nchannel = '0'+str(obj.channel)
        apl = root
        object = etree.Element('item')
        object.attrib['id'] = 'Root.Diag.DIs.' + str(obj.signal_id)[:-3] + '.ch_DI_'+nchannel
        object.attrib['value'] = val
        apl.append(object)
        tree.write(filename, pretty_print=True)

    for obj in Kzfkp.select().where(Kzfkp.typesignal == 'DO'):
        nchannel = str(obj.channel)
        val = str(obj.tag)
        if val in {'None','none'}:
            val = ' '
        if obj.channel<10:
            nchannel = '0'+str(obj.channel)
        apl = root
        object = etree.Element('item')
        object.attrib['id'] = 'Root.Diag.DOs.' + str(obj.signal_id)[:-3] + '.ch_DI_'+nchannel
        object.attrib['value'] = val
        apl.append(object)
        tree.write(filename, pretty_print=True)

def add_attrib_nameust(filename):
    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(filename, parser)
    root = tree.getroot()
    end_word = {'I':'ая',
                'T':'ая',
                'X':'ая',
                'Газ':'ая',
                'Пожар':'ая',
                'P':'ое',
                'dP':'ый',
                'L':'ый'}
    ust_name={'UstMin6':'АварийнXX минимальнXX 4',
                'UstMin5':'АварийнXX минимальнXX 3',
                'UstMin4':'АварийнXX минимальнXX 2',
                'UstMin3':'АварийнXX минимальнXX',
                'UstMin2':'МинимальнXX 2',
                'UstMin1':'МинимальнXX',
                'UstMax1':'МаксимальнXX',
                'UstMax2':'МаксимальнXX 2',
                'UstMax3':'АварийнXX максимальнXX',
                'UstMax4':'АварийнXX максимальнXX 2',
                'UstMax5':'АварийнXX максимальнXX 3',
                'UstMax6':'АварийнXX максимальнXX 4'}
    apl = root
    for obj in Analog.select():
        for ust in ust_name:
            object = etree.Element('item')
            object.attrib['id'] = 'Root.Analogs.' + str(obj.tag)+ '.Presets_IOS.'+ust
            object.attrib['value'] = str(ust_name[ust]).replace('XX', end_word[obj.sign])
            apl.append(object)
            tree.write(filename, pretty_print=True)

def add_attrib_formatanalog(filename):
    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(filename, parser)
    root = tree.getroot()
    format_analog = {'I':'1',
                'T':'1',
                'X':'1',
                'Газ':'1',
                'Пожар':'1',
                'P':'3',
                'dP':'3',
                'L':'0'}

    apl = root
    for obj in Analog.select():
        object = etree.Element('item')
        object.attrib['id'] = 'Root.Analogs.' + str(obj.tag)+ '.Format'
        object.attrib['value'] = str(format_analog[obj.sign])
        apl.append(object)
        tree.write(filename, pretty_print=True)

def add_attrib_colorshme(filename):
    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(filename, parser)
    root = tree.getroot()

    apl = root


    for obj in Discret.select():
        cs=0
            # if str_find(str(obj.description).lower(), ['двер', 'переход']):
            #     cs = 4
        if str_find(str(obj.tag).upper(), ['KKC','ED','EC','LT']):#'PC','KKC','ED','EC','LT']):
            cs = 3
        if str_find(str(obj.tag).upper(), ['ECO','ECB']):
            cs = 5
        if str_find(str(obj.tag).upper(), ['DC']):
            cs = 1
        if str_find(str(obj.tag).upper(), ['LC']):
            cs = 2





        if cs>0:
            object = etree.Element('item')
            object.attrib['id'] = "Root.Diskrets." + str(obj.tag)+ ".s_Config"
            object.attrib['value'] = str(cs)
            apl.append(object)
            tree.write(filename, pretty_print=True)




def add_map_test(filename_ua):


    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(filename_ua, parser)
    root = tree.getroot()

    try:

        apl = root
        for obj in apl.iter('node-path'):
            parent = obj.getparent()
            print(parent)
        # object = etree.Element('item')
        # object.attrib['Binding'] = 'Introduced'
        # np = etree.Element('node-path')
        # np.text = 'Root.Analogs.' + str(obj.tag) + '.AIVisualValue'
        # object.append(np)
        # namespace = etree.Element('namespace')
        # namespace.text = "nft"
        # object.append(namespace)
        # nodeType = etree.Element('nodeIdType')
        # nodeType.text = 'String'
        # object.append(nodeType)
        # nodeId = etree.Element('nodeId')
        # nodeId.text = 'AIVisualValue['+str(obj.index_arr) +']'
        # object.append(nodeId)
        # apl.append(object)




        #tree.write(filename_ua, pretty_print=True)

        print('Создан')
    except:
        print('какой то косяк')